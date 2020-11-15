#!/usr/bin/python3

import pandas as pd
#import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from sklearn import tree
from sklearn.metrics import accuracy_score
import os

def predict(name):
	print("PREDICTION",os.getcwd())
	file_path = os.getcwd()+str("/")
	#file_path = "/home/pranav/Desktop/Aura-Project/AuraSite/App/"
	file_upload = open(file_path+'Uploads/'+ str(name),'rb') 		# open the file in new data object and then load into the function
	book3 = openpyxl.load_workbook(filename=file_upload)
	sheet_name = book3.get_sheet_names()[0]

	sheet3 = book3.get_sheet_by_name(sheet_name)
	data_frame = pd.read_excel(file_upload,sheet_name = sheet_name)
	title = data_frame.columns.values[2]
	temp_names = title.split()

	person_name=""
	for i in range(3,len(temp_names)):
		person_name = person_name + str(temp_names[i]) + " "

	for i in range(4,125):
		for j in range(3,4):
			if( not sheet3.cell(row =i, column=j).value ):
				sheet3.cell(row =i, column=j).value = 0
			#print(sheet3.cell(row =i, column=j).value)

	book3.save(file_path+'Data/newwww.xlsx')
	df = pd.read_excel(file_path+'/Data/newwww.xlsx')


	selected_part = df.iloc[75:77,1:3]
	#print(selected_part.values.tolist())
	sele = selected_part.values.tolist()



	if(sele[0][0]=='Thyroid gland'):
		#print(True)
		part1 = df.iloc[1:8,0:3]
		part2 = df.iloc[11:13,0:3]
		part3 = df.iloc[14:16,0:3]
		part4 = df.iloc[17:19,0:3]
		part6 = df.iloc[22:29,0:3]
		part7 = df.iloc[31:38,0:3]
		part8 = df.iloc[40:47,0:3]
		part9 = df.iloc[48:50,0:3]
		part10 = df.iloc[53:67,0:3]
		part11 = df.iloc[70:72,0:3]
		part12 = df.iloc[76:83,0:3]
		part13 = df.iloc[85:99,0:3]
		part14 = df.iloc[100:123,0:3]
		#print(part2.values.tolist())
		#part2 = df.iloc[77:123,0:3]
		p3 = pd.concat([part1,part2,part3,part4,part6,part7,part8,part9,part10,part11,part12,part13,part14])
		p3.T.to_excel(file_path+'Data/second.xlsx')

		book_temp = openpyxl.load_workbook(file_path+'Data/second.xlsx')
		#print(book.get_sheet_names())
		sheet = book_temp.get_sheet_by_name('Sheet1')
		for i in range(4,5):
			for j in range(3,98):
				pass
				#print(sheet.cell(row=i, column=j).value)
	else:
		#print(False)

		part1 = df.iloc[1:8,0:3]
		part2 = df.iloc[11:13,0:3]
		part3 = df.iloc[14:16,0:3]
		part4 = df.iloc[17:19,0:3]
		part6 = df.iloc[22:29,0:3]
		part7 = df.iloc[31:38,0:3]
		part8 = df.iloc[40:47,0:3]
		part9 = df.iloc[48:50,0:3]
		part10 = df.iloc[53:67,0:3]
		part11 = df.iloc[70:72,0:3]
		part12 = df.iloc[74:81,0:3]
		part13 = df.iloc[83:97,0:3]
		part14 = df.iloc[98:121,0:3]
		#print(part2.values.tolist())
		#part2 = df.iloc[77:123,0:3]
		p3 = pd.concat([part1,part2,part3,part4,part6,part7,part8,part9,part10,part11,part12,part13,part14])
		p3.T.to_excel(file_path+'Data/second.xlsx')

		book_temp = openpyxl.load_workbook(file_path+'Data/second.xlsx')
		#print(book.get_sheet_names())
		sheet = book_temp.get_sheet_by_name('Sheet1')
	'''
		for i in range(4,5):
			for j in range(3,98):
				print(sheet.cell(row=i, column=j).value)
	'''


	file = file_path+"Data/non_meditator_data.xlsx"
	book = openpyxl.load_workbook(file)
	#print(book.get_sheet_names())
	sheet = book.get_sheet_by_name('Sheet1')

	file1 = file_path+'Data/meditator_data.xlsx'
	book1 = openpyxl.load_workbook(file1)
	#print(book.get_sheet_names())
	sheet1 = book1.get_sheet_by_name('Sheet1')

	file2 = file_path+"Data/second.xlsx"
	book2 = openpyxl.load_workbook(file2)
	#print(book.get_sheet_names())
	sheet2 = book2.get_sheet_by_name('Sheet1')


	# for training data
	train_rows=[]
	train_columns=[]

	# for non_meditators
	for i in range(5,11):				# range for the non_meditation is 5 to 56
		row=[]
		for j in range(4,99):
			row.append(sheet.cell(row=i, column=j).value)
		train_rows.append(row)
		train_columns.append(sheet.cell(row=i, column=2).value)


	# for meditators
	for i in range(5,22):				#  range for meditator is 5 to 24
		row=[]
		for j in range(4,99):
			row.append(sheet.cell(row=i, column=j).value)
		train_rows.append(row)
		train_columns.append(sheet1.cell(row=i, column=2).value)

	# for testing

	test_rows=[]
	test_columns=[]

	for i in range(4,5):
		row=[]
		for j in range(3,98):
			row.append(sheet2.cell(row=i, column=j).value)
		test_rows.append(row)

	#print(test_rows)
	# Decision tree
	algo = tree.DecisionTreeClassifier()
	trained = algo.fit(X = train_rows,y = train_columns)
	result = trained.predict(test_rows)
	int_result = int(result[0])

	#return the result as boolean (True for meditator , False for non meditator)
	if (int_result):
		return {"name":person_name,"status":True}
	else:
		return {"name":person_name,"status":False}
