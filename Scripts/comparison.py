#!/usr/bin/python3

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import gridspec
import openpyxl
import numpy as np

def compare(file1,file2):

	#READING IDEAL FILE
	file_ideal = '/home/pranav/Desktop/Aura-Project/AuraSite/App/Data/ideal_mukesh.xlsx'		#CHANGE PATH
	file1='/home/pranav/Desktop/Aura-Project/AuraSite/App/Uploads/'+file1				#CHANGE PATH
	file2='/home/pranav/Desktop/Aura-Project/AuraSite/App/Uploads/'+file2				#CHANGE PATH
	file1=open(file1,'rb')
	file2=open(file2,'rb')
	#READING BOOK AND SHEET NAME
	book = openpyxl.load_workbook(file_ideal)
	name = book.get_sheet_names()[0]

	book1 = openpyxl.load_workbook(file1)
	name1 = book1.get_sheet_names()[0]

	book2 = openpyxl.load_workbook(file2)
	name2 = book2.get_sheet_names()[0]

	#FEATCHING COMPLETE SHEET DATA
	data = pd.read_excel(file_ideal, sheet_name=name)
	data1 = pd.read_excel(file1, sheet_name=name1)
	data2 = pd.read_excel(file2, sheet_name=name2)

	#EXTRACTING DATA INTO LIST FOR GRAPH1 (GENERAL)
	plot_values = data.iloc[3:6,1:3]
	plot_list = plot_values.values.tolist()

	plot_values1 = data1.iloc[3:6,1:3]
	plot_list1 = plot_values1.values.tolist()

	plot_values2 = data2.iloc[3:6,1:3]
	plot_list2 = plot_values2.values.tolist()


	#EXTRACTING DATA INTO LIST FOR GRAPH2 (CHAKRA)
	plot_values_chakra1 = data1.iloc[22:29,1:3]
	plot_list_chakra1 = plot_values_chakra1.values.tolist()

	plot_values_chakra2 = data2.iloc[22:29,1:3]
	plot_list_chakra2 = plot_values_chakra2.values.tolist()


	#SEPERATING LIST DATA FOR CO-ORDINATES (GENERAL)
	x_axis_ideal=[]
	y_axis_ideal=[]
	y_axis_file1=[]
	y_axis_file2=[]

	for i in range(0,len(plot_list)):
		x_axis_ideal.append(plot_list[i][0])
		y_axis_ideal.append(plot_list[i][1])
		y_axis_file1.append(plot_list1[i][1])
		y_axis_file2.append(plot_list2[i][1])


	#SEPERATING LIST DATA FOR CO-ORDINATES (CHAKRA)
	x_chakra=[]
	y_chakra_file1=[]
	y_chakra_file2=[]

	for i in range(0,len(plot_list_chakra1)):
		x_chakra.append(plot_list_chakra1[i][0])
		y_chakra_file1.append(plot_list_chakra1[i][1])
		y_chakra_file2.append(plot_list_chakra2[i][1])


	#FORMATION GENERAL GRAPH ----
	#CREATING PLOT ARGUMENTS
	N = 3
	ind = np.arange(N)
	width = 0.2

	#CREATING SUB-POLTS
	fig = plt.figure(figsize=(10,6))
	gs = gridspec.GridSpec(1, 1)
	graph = plt.subplot(gs[0])
	p1 = graph.bar( ind-width, y_axis_file1, width=0.2, color='b', align='center')
	p2 = graph.bar( ind      , y_axis_ideal, width=0.2, color='g', align='center')
	p3 = graph.bar( ind+width, y_axis_file2, width=0.2, color='r', align='center')

	#PLOT ATTRIBUTES
	graph.set_title('Comparitive Analysis')

	graph.set_xticks(ind + (width/999) / 2)
	graph.set_xticklabels(x_axis_ideal)

	graph.legend((p1[0], p2[0], p3[0]), ('User 1', 'Ideal', 'User 2'))
	graph.autoscale_view()

	plt.savefig('/home/pranav/Desktop/Aura-Project/AuraSite/App/static/comparison_img/general_graph.png')		#CHANGE PATH
	plt.clf()


	graph1_path='/home/pranav/Desktop/Aura-Project/AuraSite/App/static/comparison_img/general_graph.png'

	#FORMATION CHAKRA GRAPH ----
	#CREATING PLOT ARGUMENTS
	N = 7
	ind = np.arange(N)
	width = 0.3

	#CREATING SUB-POLTS
	graph = plt.subplot(gs[0])
	g1 = graph.bar( ind      , y_chakra_file1, width=0.3, color='b', align='center')
	g2 = graph.bar( ind+width, y_chakra_file2, width=0.3, color='g', align='center')

	#PLOT ATTRIBUTES
	graph.set_title('Chakra Analysis')

	graph.set_xticks(ind + width / 2)
	graph.set_xticklabels(x_chakra,rotation=20)
	graph.legend((g1[0], g2[0]), ('User 1', 'User 2'))
	graph.autoscale_view()

	plt.savefig('/home/pranav/Desktop/Aura-Project/AuraSite/App/static/comparison_img/chakra_graph.png')		#CHANGE PATH
	graph2_path='/home/pranav/Desktop/Aura-Project/AuraSite/App/static/comparison_img/chakra_graph.png'

	graph_path=[graph1_path,graph2_path]

	return graph_path
	'''
	#NEW UPDATED REQUIREMENT
	#FOR FILE1
	general_file1_part1 = data1.iloc[2:8,1:3]
	general_file1_part2 = data1.iloc[53:55,1:3]
	general_list1_part1 = general_file1_part1.values.tolist()
	general_list1_part2 = general_file1_part2.values.tolist()

	general_list1_part1.append(general_list1_part2)
	finalval_file1=general_list1_part1

	#print(data)
	#print(finalval_file1)

	#FOR FILE2
	general_file2_part1 = data2.iloc[2:8,1:3]
	general_file2_part2 = data2.iloc[53:55,1:3]
	general_list2_part1 = general_file2_part1.values.tolist()
	general_list2_part2 = general_file2_part2.values.tolist()

	general_list2_part1.append(general_list2_part2)
	finalval_file2=general_list2_part1

	#print(data)
	#print(finalval_file2)

	return finalval_file1,finalval_file2



	#COMPARING RESULTS WITH IDEAL
	ideal_values = data.iloc[2:5,1:3]
	ideal_list = ideal_values.values.tolist()
	#print(ideal_list)

	file1_values = data1.iloc[2:5,1:3]
	file1_list = file1_values.values.tolist()
	#print(file1_list)

	file2_values = data2.iloc[2:5,1:3]
	file2_list = file2_values.values.tolist()
	#print(file2_list)

	file1_dict={'Emotional Pressure':'None','Energy':'None','L/R Symmetry':'None'}
	file2_dict={'Emotional Pressure':'None','Energy':'None','L/R Symmetry':'None'}


	#CHECKING EMOTIONAL PRESSURE (WINDOW 0.5)
	emo_differ_f1=(ideal_list[0][1]-file1_list[0][1])
	emo_differ_f2=(ideal_list[0][1]-file2_list[0][1])

	#FOR FILE1
	if(emo_differ_f1 < -0.5):
		file1_dict['Emotional Pressure']='High'
	elif(emo_differ_f1 > 0.5):
		file1_dict['Emotional Pressure']='Low'
	else:
		file1_dict['Emotional Pressure']='Normal'

	#FOR FILE2
	if(emo_differ_f2 < -0.5):
		file2_dict['Emotional Pressure']='High'
	elif(emo_differ_f2 > 0.5):
		file2_dict['Emotional Pressure']='Low'
	else:
		file2_dict['Emotional Pressure']='Normal'


	#CHECKING ENERGY (WINDOW 15)
	energy_differ_f1=(ideal_list[1][1]-file1_list[1][1])
	energy_differ_f2=(ideal_list[1][1]-file2_list[1][1])

	#FOR FILE1
	if(energy_differ_f1 < -15):
		file1_dict['Energy']='High'
	elif(energy_differ_f1 > 15):
		file1_dict['Energy']='Low'
	else:
		file1_dict['Energy']='Normal'

	#FOR FILE2
	if(energy_differ_f2 < -15):
		file2_dict['Energy']='High'
	elif(energy_differ_f2 > 15):
		file2_dict['Energy']='Low'
	else:
		file2_dict['Energy']='Normal'


	#CHECKING L/R Symmetry (WINDOW 5)
	symm_differ_f1=(ideal_list[2][1]-file1_list[2][1])
	symm_differ_f2=(ideal_list[2][1]-file2_list[2][1])

	#FOR FILE1
	if(symm_differ_f1 < -15):
		file1_dict['L/R Symmetry']='High'
	elif(symm_differ_f1 > 15):
		file1_dict['L/R Symmetry']='Low'
	else:
		file1_dict['L/R Symmetry']='Normal'

	#FOR FILE2
	if(symm_differ_f2 < -5):
		file2_dict['L/R Symmetry']='High'
	elif(symm_differ_f2 > 5):
		file2_dict['L/R Symmetry']='Low'
	else:
		file2_dict['L/R Symmetry']='Normal'

	return file1_dict,file2_dict
	'''
